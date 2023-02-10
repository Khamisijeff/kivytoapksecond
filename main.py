
import jdatetime
from bidi.algorithm import get_display
from arabic_reshaper import reshape
from jdatetime import datetime
import pyperclip
import sqlite3
from shutil import copy2
from random import choice
import openpyxl
from os.path import exists
import shutil
from time import sleep


from kivy.animation import Animation
from kivy.properties import  StringProperty,NumericProperty
from kivy.clock import Clock
from kivy.properties import DictProperty
from kivy.uix.screenmanager import ScreenManager, Screen, NoTransition
from kivy.metrics import dp
from kivy.lang import Builder

from kivymd.uix.button import MDRaisedButton
from kivymd.app import MDApp
from kivymd.uix.floatlayout import MDFloatLayout
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.toast import toast
from kivymd.uix.card import MDCard
from kivymd.uix.tooltip import MDTooltip
from kivymd.uix.behaviors import MagicBehavior,ScaleBehavior
from kivymd.uix.behaviors.rotate_behavior import RotateBehavior
from kivymd.uix.dialog import MDDialog
from kivymd.color_definitions import colors
from kivymd.uix.button import MDRoundFlatButton
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.circularlayout import MDCircularLayout
from kivymd.uix.behaviors.elevation import CommonElevationBehavior
from kivymd.uix.list import TwoLineRightIconListItem,OneLineAvatarIconListItem,TwoLineListItem,ThreeLineListItem
from kivymd.uix.datatables import MDDataTable



Builder.load_file('startscreen.kv')
from kivy.core.window import Window
Window.size=(350,600)



class Mainscreenunits(MDFloatLayout,CommonElevationBehavior):
    nametxt=''
    pad_color=''
    phone=''
    unit_number=''
    enterance_date=''

    def getunitreportpayment(self,unitnumber):
        persian_numbers=['۰','۱','۲','۳','۴','۵','۶','۷','۸','۹','۱۰','۱۱','۱۲','۱۳','۱۴','۱۵','۱۶','۱۷','۱۸','۱۹','۰','۲۰','۲۱','۲۲','۲۳','۲۴','۲۵','۲۶','۲۷','۲۸','۲۹','۳۰']
        eng_unit_number=NumericProperty(0)
        total_unit_numbers=NumericProperty(0)
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        c.execute(""" SELECT * FROM setting """)
        if unitnumber in persian_numbers:
            eng_unit_number=persian_numbers.index(unitnumber)

            # Get the total number of units from setting table:
            total = c.fetchall()
            for i in total:
                total_unit_numbers = int(i[3])

            # connect to related Unit Number table:
            if eng_unit_number<=total_unit_numbers:
                c.execute(f""" SELECT * FROM unit{eng_unit_number}""")
                report = c.fetchall()
                if len(report) == 0:
                    Mainapp().system_message('. همسایه ما واریزی نداشته ', 'باشه')
                else:
                    self.raise_Home_Unit_Payment_Report()
                    sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_x = 1
                    sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_y = 1
                    sm.get_screen('main').ids.Home_Unit_Payment_Report.elevation = 3
                    for i in report:
                        Home_Unit_Payment_Report_Card.report_date = i[2]
                        Home_Unit_Payment_Report_Card.report_amount = i[1]
                        Home_Unit_Payment_Report_Card.report_for = i[0]
                        sm.get_screen('main').total_pay+=int(i[1])
                        sm.get_screen('main').ids.home_report_payment.add_widget(Home_Unit_Payment_Report_Card())
            else:
                Mainapp().system_message(f'. ساختمان {total_unit_numbers} واحد دارد ', 'باشه')

        else:
            # Check if the entered unit number coincides the total building unit numbers Get:
            eng_unit_number=unitNumber
            c.execute(""" SELECT * FROM setting """)
            total = c.fetchall()
            total_unit_numbers = int(i[3])
            # connect to related Unit Number table:
            c.execute(f""" SELECT * FROM unit{eng_unit_number}""")
            report = c.fetchall()
            if len(report) == 0:
                Mainapp().system_message('. همسایه ما واریزی نداشته ', 'باشه')
            else:
                self.raise_Home_Unit_Payment_Report()
                sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_x = 1
                sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_y = 1
                sm.get_screen('main').ids.Home_Unit_Payment_Report.elevation = 3
                for i in report:
                    Home_Unit_Payment_Report_Card.report_date = i[2]
                    Home_Unit_Payment_Report_Card.report_amount = i[1]
                    Home_Unit_Payment_Report_Card.report_for = i[0]
                    sm.get_screen('main').total_pay += int(i[1])
                    sm.get_screen('main').ids.home_report_payment.add_widget(Home_Unit_Payment_Report_Card())








        conn.commit()
        conn.close()


    def raise_Home_Unit_Payment_Report(self):
        Animation(pos_hint={'center_y':0.5},d=0.3).start(sm.get_screen('main').ids.Home_Unit_Payment_Report)
        print('done')
    def getunitphonenumber(self,neighbor_number):
        Mainapp().system_message(f'{neighbor_number}', 'عالیه')

    def get_PaymentList(self):
        sm.get_screen('main').showaddincome()

class AddNote(MDFloatLayout):
    def addtasktodatabase(self, *args):
        if self.ids.task_title_txt.text!='' and self.ids.task_desc_txt.text!='':
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute("""INSERT INTO new_tasks VALUES (:ititle,:idescription) """,
                      {'ititle': self.ids.task_title_txt.text,
                       'idescription': self.ids.task_desc_txt.text,
                       })
            conn.commit()
            conn.close()

            sm.get_screen('main').get_task_number()
            sm.get_screen('main').retrievedonenotes()
            sm.get_screen('main').retrieveundonenotes()

            self.ids.task_title_txt.text=''
            self.ids.task_title_input.text=''
            self.ids.task_desc_txt.text=''
            self.ids.task_desc_input.text=''
            Mainapp().system_message('.یادداشت جدید افزوده شد', 'عالیه')
        else:
            Mainapp().system_message('.عنوان و شرح را وارد کنید', 'باشه')

class AddIncome(MDFloatLayout):
    def get_date(self):
        self.ids.add_income_date_text.text = MyScreen().get_date()

        # Add new income
    def add_new_income(self, *args):
        eng_unit_number=StringProperty('0')
        total_unit_numbers=NumericProperty(0)
        persian_numbers = ['۰', '۱', '۲', '۳', '۴', '۵', '۶', '۷', '۸', '۹', '۱۰', '۱۱', '۱۲', '۱۳', '۱۴', '۱۵',
                           '۱۶',
                           '۱۷', '۱۸', '۱۹', '۲۰', '۲۱', '۲۲', '۲۳', '۲۴', '۲۵', '۲۶', '۲۷', '۲۸', '۲۹',
                           '۳۰']
        if self.ids.add_unit_text.text and self.ids.add_income_date_text.text and self.ids.add_amount_text.text and self.ids.add_for_text.text != '':
            # create cursor

            conn = sqlite3.connect('database.db')
            c = conn.cursor()

            # Check if the entered unit number coincides the total building unit numbers Get:
            c.execute(""" SELECT * FROM setting """)
            total=c.fetchall()
            for i in total:
                total_unit_numbers=int(i[3])

            if self.ids.add_unit_text.text in persian_numbers:
                eng_unit_number = persian_numbers.index(self.ids.add_unit_text.text)
            else:
                eng_unit_number=self.ids.add_unit_text.text

            if int(eng_unit_number)<=total_unit_numbers:
                # insert into table
                c.execute(
                    """INSERT INTO incomes VALUES(:unit, :idate, :amount, :ifor) """,
                    {
                        'unit':eng_unit_number ,
                        'idate': self.ids.add_income_date_text.text,
                        'amount': self.ids.add_amount_text.text,
                        'ifor': self.ids.add_for_text.text,

                    })
                c.execute(
                    f"""INSERT INTO unit{eng_unit_number} VALUES(:idate, :amount, :ifor) """,
                    {

                        'idate': self.ids.add_income_date_text.text,
                        'amount': self.ids.add_amount_text.text,
                        'ifor': self.ids.add_for_text.text,

                        })

                # Message that function has been Done.
                Mainapp().system_message('.واریزی جدید افزوده شد', 'عالیه')
                sm.get_screen('main').Retrieve_from_income_table()
            else:
                Mainapp().system_message('.واحد وجود ندارد', 'باشه')

            # Clear all the Fields:
            self.ids.add_unit_text.text = ''
            self.ids.add_income_date_text.text = ''
            self.ids.add_amount_text.text = ''
            self.ids.add_for_text.text = ''
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')
        conn.commit()
        conn.close()


class AddPayment(MDFloatLayout):
    def get_date(self):
        self.ids.p_date_text.text=MyScreen().get_date()
    def add_new_payment(self, *args):
        if self.ids.p_date_text.text and self.ids.p_amount_text.text and self.ids.p_for_text.text and self.ids.whom_text.text !='':
            conn = sqlite3.connect('database.db')
            # create cursor
            c = conn.cursor()

            # insert into table
            c.execute(
                """INSERT INTO payments VALUES(:idate, :amount, :ifor, :whom) """,
                {
                    'idate': self.ids.p_date_text.text,
                    'amount': self.ids.p_amount_text.text,
                    'ifor': self.ids.p_for_text.text,
                    'whom': self.ids.whom_text.text,

                })
            conn.commit()
            conn.close()

            # Clear all the Fields:
            self.ids.p_date_text.text = ''
            self.ids.p_amount_text.text = ''
            self.ids.p_for_text.text = ''
            self.ids.whom_text.text = ''

            # Message that function has been Done.
            Mainapp().system_message('. پرداختی افزوده شد', 'عالیه')
            sm.get_screen('main').Retrieve_from_payments_table()


        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')



class AddBuildingCog(MDFloatLayout):
    def add_new_building_cog(self, *args):
        if self.ids.b_name_text.text and self.ids.b_type_text.text and self.ids.b_tel_text.text != '':
            conn = sqlite3.connect('database.db')
            # create cursor
            c = conn.cursor()

            # insert into table
            c.execute(
                """INSERT INTO building_cog VALUES (:iname, :itype, :itel) """,
                {
                    'iname': self.ids.b_name_text.text,
                    'itype': self.ids.b_type_text.text,
                    'itel': self.ids.b_tel_text.text,

                })

            conn.commit()
            conn.close()
            # Clear all the Fields:
            self.ids.b_name_text.text = ''
            self.ids.b_type_text.text = ''
            self.ids.b_tel_text.text = ''


            # Message that function has been Done.
            Mainapp().system_message('.خدمات جدید افزوده شد', 'عالیه')
            sm.get_screen('main').Retrieve_from_building_cogs_table()
            sm.get_screen('main').get_service_number()

        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')


class AddResident(MDFloatLayout):
    def addresidenttodatabase(self):
        if self.ids.resident_name_txt.text and self.ids.resident_family_name_txt.text and self.ids.resident_phone_txt.text and self.ids.resident_unit_txt.text and self.ids.resident_date_in_txt and self.ids.resident_family_number_txt != '':
            conn = sqlite3.connect('database.db')
            # create cursor
            c = conn.cursor()
            # insert into table
            c.execute(
                """INSERT INTO neighbors VALUES(:iName,:FamilyName,:PhoneNumber,:UnitNumber,:Member,:DateIn) """,
                {
                    'iName': self.ids.resident_name_txt.text,
                    'FamilyName': self.ids.resident_family_name_txt.text,
                    'PhoneNumber': self.ids.resident_phone_txt.text,
                    'UnitNumber': self.ids.resident_unit_txt.text,
                    'Member': self.ids.resident_family_number_txt.text,
                    'DateIn': self.ids.resident_date_in_txt.text,

                })
            conn.commit()
            conn.close()
            # Clear all the Fields:
            self.ids.resident_name_txt.text = ''
            self.ids.resident_unit_txt.text = ''
            self.ids.resident_family_name_txt.text = ''
            self.ids.resident_phone_txt.text = ''
            self.ids.resident_family_number_txt.text = ''
            self.ids.resident_date_in_txt.text = ''


            # Message that function has been Done.
            Mainapp().system_message('.همسایه جدید افزوده شد', 'عالیه')
            sm.get_screen('main').get_hamsaye_number()
            sm.get_screen('main').addunitnumberstomain()
            sm.get_screen('main').Retrieve_from_neighbors_list()
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')
    # Create a new account
class AddOwner(MDFloatLayout):
    def addownertodatabase(self, *args):
        if self.ids.owner_name_txt.text!='' and self.ids.owner_family_name_txt.text!='' and self.ids.owner_phone_txt.text!='' and self.ids.owner_unit_txt.text!='':
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute("""INSERT INTO owners VALUES (:iname,:ifamily,:iphone,:iunit) """,
                      {'iname': self.ids.owner_name_txt.text,
                       'ifamily': self.ids.owner_family_name_txt.text,
                       'iphone': self.ids.owner_phone_txt.text,
                       'iunit': self.ids.owner_unit_txt.text,
                       })
            conn.commit()
            conn.close()


            self.ids.owner_name_txt.text=''
            self.ids.owner_name_input.text=''
            self.ids.owner_family_name_txt.text = ''
            self.ids.owner_family_name_txt.text = ''
            self.ids.owner_unit_txt.text = ''
            self.ids.owner_unit_input.text = ''
            self.ids.owner_phone_txt.text = ''
            self.ids.owner_phone_input.text = ''
            Mainapp().system_message('.مالک جدید افزوده شد', 'عالیه')
            sm.get_screen('main').get_owners_number()
            sm.get_screen('main').Retrieve_from_owners_list()

        else:
            Mainapp().system_message('.عنوان و شرح را وارد کنید', 'باشه')













class StartScreen(Screen):
    # This will show and hide a dialogbox from the bottom of the app
    # which help you to register the building, it's unit numbers and the manager name

    show_addbuilding = 0
    def showaddbuilding(self, *args):
        if self.show_addbuilding == 0:
            anim = Animation(pos_hint={'center_x': 0.5, 'center_y': 0.35}, d=0.5)
            anim.start(self.ids.add_building)
            self.show_addbuilding = 1
        else:
            Animation(pos_hint={'center_x': 0.5, 'center_y': -1}, d=0.5).start(self.ids.add_building)
            self.show_addbuilding = 0


    def createbuilding(self,building_name,manager_name,total_unit_number):
        if self.ids.building_name_txt.text and self.ids.building_manager_name_txt.text and self.ids.building_unit_number_input.text !='':
            if not exists('database.db'):
                # connect to database
                import sqlite3
                conn = sqlite3.connect('database.db')
                # create cursor
                c = conn.cursor()

                # create tables

                c.execute(
                    """ CREATE TABLE owners ("iName" text,"FamilyName" text ,"PhoneNumber" text, 'UnitNumber' text) """)
                # Create a table for all users this deleted neighbors are in this table
                c.execute(
                    """ CREATE TABLE neighbors ("iName" text,"FamilyName" text ,"PhoneNumber" text, "DateIn" text, 'UnitNumber' text,'FamilyMember' text) """)
                # Create a table for all users this deleted neighbors are in this table

                c.execute(
                    """ CREATE TABLE deleted_neighbors ("iName" text,"FamilyName" text ,"PhoneNumber" text, "DateIn" text, 'UnitNumber' text,'FamilyMember' text) """)
                # Where incomes registered
                c.execute(
                    """CREATE TABLE incomes ("unit_text" text,"income_date" text, "amount" text, "for_info" text)""")

                # Where Payments registered
                c.execute(
                    """CREATE TABLE payments ("payment_date" text, "amount" text, "for_info" text,"to_whom" text )""")

                # Where Building Cogs registered
                c.execute(
                    """CREATE TABLE building_cog ("name" text, "type" text, "phone" text)""")
                # Where Tasks are registered
                c.execute(
                    """CREATE TABLE new_tasks ("title" text, "description" text)""")
                c.execute(
                    """CREATE TABLE done_tasks ("title" text, "description" text)""")
                # Where Settings registered
                c.execute(
                    """CREATE TABLE setting ("color" text, "ManagerName" text,"BuildingName" text,'UnitNumbers' text)""")
                c.execute(""" INSERT INTO setting VALUES(:my_color,:manager_name,:building_name,:unit_number)""",
                          {'my_color': 'Blue',
                           'manager_name': building_name ,
                           'building_name':manager_name,
                           'unit_number':total_unit_number

                           })
                for i in range(1,int(total_unit_number)+1):
                    c.execute(F"""CREATE TABLE unit{i}('Date' text,'Amount' text,'For' text) """)
                conn.commit()
                conn.close()
                sm.current='main'
            else:
                pass
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')
class Donenote(ThreeLineListItem):
    title=''
    description=''

    def remove_note(self,title):
        sm.get_screen('main').ids.donenotepad.remove_widget(self)
        conn = sqlite3.connect('database.db')
        c = conn.cursor()

        c.execute(f""" DELETE FROM done_tasks WHERE title=:jtitle""",
                  {'jtitle': title})
        conn.commit()
        conn.close()

class MyNote(MDFloatLayout):
    title=''
    description=''
    def checked(self,instance,value,title,desc):
        if value:
            self.ids.note.text=f'[s]{self.ids.note.text}[/s]'
            # ADD to done pad and done table:
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute("""INSERT INTO done_tasks VALUES (:ititle,:idescription) """,
                      {'ititle': title,
                       'idescription': desc,
                       })
            c.execute(f""" DELETE FROM new_tasks WHERE title=:jtitle""",
                      {'jtitle':title})
            conn.commit()
            conn.close()
            Donenote.title=title
            Donenote.description=desc
            sm.get_screen('main').ids.donenotepad.add_widget(Donenote())
            sm.get_screen('main').ids.undonenotepad.remove_widget(self)
            sm.get_screen('main').on_enter()

        #else:
         #   self.ids.note.text = self.ids.note.text[3:len(self.ids.note.text)-4]


class My_icon_tip(MDFloatLayout,MDTooltip):
    pass

class My_MDFloatingActionButtonSpeedDial(MDFloatLayout):
    pass

class Mycolors(MDRoundFlatButton, MagicBehavior):
    pass
class MyCircularLayout(MDCircularLayout):
    pass
class IncomesListReportItems(MDFloatLayout):
    unit_text=''
    date_text=''
    amount_text=''
    for_text=''
class PaymentsListReportItems(MDFloatLayout):
    date_text=''
    amount_text=''
    for_text=''
    whom_text=''
class BuildingCogListReportItems(MDFloatLayout,MDTooltip):
    name_text=''
    type_text=''
    phone_text=StringProperty("")
    bg_color=''
    def clicked(self,itext):
        pyperclip.copy(f'{itext}')
        toast('Ok')



class NeighborsListReportItems(MDFloatLayout):
    my_text=''
    secondary_text=''

    def remove_account(self, meyar, *args):
        Mainapp().system_message('.همسایه حذف شد', 'باشه ')
        sm.get_screen('main').ids.neighbor_pad.remove_widget(self)

        # connect to database
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        # c.execute(""" SELECT * FROM neighbors""")
        c.execute(""" DELETE FROM neighbors WHERE PhoneNumber=:meyar""", {
            'meyar': meyar
        })
        # Applying Changes
        conn.commit()
        print('Done')
        conn.close()
        sleep(1)
        sm.get_screen('main').get_hamsaye_number()
        sm.get_screen('main').Retrieve_from_neighbors_list()
class OwnersListReportItems(MDFloatLayout):
    my_text =StringProperty('')
    secondary_text = StringProperty('')


    def show_info(self,phone):
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        owners_data=c.execute("""SELECT * FROM owners WHERE PhoneNumber=:phone""",{
            'phone':phone
        })
        for i in owners_data:
            print(i)


        conn.commit()
        conn.close()

    def remove_owner(self,imeyar,*args):
        Mainapp().system_message('.مالک حذف شد','باشه ')
        sm.get_screen('main').ids.owners_pad.remove_widget(self)

        # connect to database
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        #c.execute(""" SELECT * FROM neighbors""")
        c.execute(""" DELETE FROM owners WHERE PhoneNumber=:meyar""" ,{
            'meyar': imeyar
        })
        # Applying Changes
        conn.commit()
        print('Done')
        conn.close()
        sleep(1)
        sm.get_screen('main').get_owners_number()
        sm.get_screen('main').Retrieve_from_owners_list()

class LoginContent(Screen):
   pass

class ContentNavigationDrawer(MDBoxLayout):
    pass

class Add_neighbor_content(MDBoxLayout):
    pass

class ItemConfirm(OneLineAvatarIconListItem):

    text=''
    font_name='bijan.ttf'


class Neighbors_Content(MDBoxLayout):
    pass
class Home_Unit_Payment_Report_Card(MDCard):
    report_date=StringProperty('')
    report_amount=StringProperty('')
    report_for=StringProperty('')



class Home_Unit_Payment_Report(MDCard,ScaleBehavior):
    pay_date=StringProperty('')
    pay_amount=StringProperty('')
    pay_for=StringProperty('')
    scale_value_x=NumericProperty(0)
    scale_value_y= NumericProperty(0)
    def closebox(self):
        sm.get_screen('main').ids.home_report_payment.clear_widgets()
        sm.get_screen('main').ids.Home_Unit_Payment_Report.elevation=0
        Animation(pos_hint={'center_y':-1},d=0.5).start(sm.get_screen('main').ids.Home_Unit_Payment_Report)
        sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_x=0
        sm.get_screen('main').ids.Home_Unit_Payment_Report.scale_value_y=0
        sm.get_screen('main').total_pay=0





class MyScreen(Screen):
    total_pay = NumericProperty('0')

    def addunitnumberstomain(self):
        bg_colors = ['Red', 'Pink', 'Purple', 'DeepPurple', 'Indigo', 'Blue', 'LightBlue', 'Cyan', 'Teal', 'Green', 'LightGreen', 'Lime', 'Amber', 'Orange', 'DeepOrange', 'Brown', 'Gray', 'BlueGray']
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        c.execute("SELECT * FROM neighbors")
        Neighbors=c.fetchall()
        self.ids.mainscreennumberpad.clear_widgets()
        for neighbor in Neighbors:
            Mainscreenunits.pad_color=colors[choice(bg_colors)]['500']
            Mainscreenunits.nametxt=(reshape(f'{neighbor[1]} {neighbor[0]}'))
            Mainscreenunits.unit_number=f'{neighbor[3]}'
            Mainscreenunits.enterance_date=f'{neighbor[5]}'
            Mainscreenunits.phone=f'{neighbor[2]}'
            self.ids.mainscreennumberpad.add_widget(Mainscreenunits())

        conn.commit()
        conn.close()



    def backup_database(self):
        self.file_browser=MDFileManager(
            exit_manager=self.exit_manager,
            select_path=self.select_back_up_path,
            preview=True
        )
        self.file_browser.show('/')


    def start_browsing(self):
        self.file_browser = MDFileManager(
            exit_manager=self.exit_manager,
            select_path=self.select_path,
        )
        self.file_browser.show('/')

    def exit_manager(self,*args):
        try:
            self.file_browser.close()
            self.change_screen('main')
        except TypeError:
            pass

    def select_path(self, path):
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        book = openpyxl.Workbook()
        income_sheet = book.create_sheet('واریزی')
        payment_sheet = book.create_sheet('پرداختی')
        done_task_sheet = book.create_sheet('یادآوری انجام شده ')
        new_task_sheet = book.create_sheet('یادآوری انجام نشده ')
        building_cog_sheet = book.create_sheet('خدمات')
        now = jdatetime.datetime.now()
        now_time = now.strftime("%d.%m.%y %H:%M:%S")
        i = 0
        c.execute(""" SELECT * FROM incomes""")
        results = c.fetchall()
        for row in results:
            i += 1
            j = 1
            for col in row:
                cell = income_sheet.cell(row=i, column=j)
                cell.value = get_display(col)
                j += 1
        i = 0
        c.execute(""" SELECT * FROM payments""")
        results = c.fetchall()
        for row in results:
            i += 1
            j = 1
            for col in row:
                cell = payment_sheet.cell(row=i, column=j)
                cell.value = get_display(col)
                j += 1
        i = 0
        c.execute(""" SELECT * FROM done_tasks""")
        results = c.fetchall()
        for row in results:
            i += 1
            j = 1
            for col in row:
                cell = done_task_sheet.cell(row=i, column=j)
                cell.value = get_display(col)
                j += 1

        i = 0
        c.execute(""" SELECT * FROM new_tasks""")
        results = c.fetchall()
        for row in results:
            i += 1
            j = 1
            for col in row:
                cell = new_task_sheet.cell(row=i, column=j)
                cell.value = get_display(col)
                j += 1
        i = 0
        c.execute(""" SELECT * FROM building_cog""")
        results = c.fetchall()
        for row in results:
            i += 1
            j = 1
            for col in row:
                cell = building_cog_sheet.cell(row=i, column=j)
                cell.value = get_display(col)
                j += 1
        book.save(f'{path}/result{now_time}.xls')
        self.change_screen('main')
        self.exit_manager()
        conn.commit()
        conn.close()

    def get_task_number(self):
        bengesht=[]
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        hamsayeha=c.execute("""SELECT * FROM new_tasks""")
        for i in hamsayeha:
            bengesht.append(i)
        self.ids.task.right_text=f'{len(bengesht)}'

    def get_hamsaye_number(self):
        goose=[]
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        hamsayeha=c.execute("""SELECT * FROM neighbors""")
        for i in hamsayeha:
            goose.append(i)
        self.ids.hamsaye.right_text=f'{len(goose)}'

    def get_owners_number(self):
        ownerslist=[]
        conn=sqlite3.connect('database.db')
        c=conn.cursor()
        c.execute("""SELECT * FROM owners""")
        owners=c.fetchall()
        for owner in owners:
            ownerslist.append(owner)
        self.ids.owners.right_text=f'{len(ownerslist)}'

    def get_service_number(self):
        duck = []
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        services = c.execute("""SELECT * FROM building_cog""")
        for i in services:
            duck.append(i)
        self.ids.service.right_text = f'{len(duck)}'







    def get_date(self):
        self.sal=jdatetime.datetime.now().year
        self.mah=jdatetime.datetime.now().month
        self.rooz=jdatetime.datetime.now().day
        return (f'{self.sal}.{self.mah}.{self.rooz}')

    time_hour=StringProperty('00')
    time_minute=StringProperty('00')

    def get_time(self,*args):
        hour=f'{jdatetime.datetime.now().hour}'
        minute=f'{jdatetime.datetime.now().minute}'
        if int(hour) <=9:
            self.time_hour=f'0{hour}'
        else:
            self.time_hour=f'{hour}'
        if int(minute)<=9:
            self.time_minute=f'0{minute}'
        else:
            self.time_minute=f'{minute}'



    data=DictProperty()

    def speeddialStart(self):

        # if self.show==0:
        #     pass
        # else:
        #     self.showaddnote()



        self.data={
            "1":['cash-plus',"on_press",lambda x:self.showaddincome()],
            "2":['cash-minus',"on_press",lambda x:self.showaddpayment()],
            "3":['office-building-cog',"on_press",lambda x:self.showaddbuildingcog()],
            "4": ['note-edit', "on_press", lambda x: self.showaddnotemain()],
            "5":['account','on_press',lambda x: self.showaddowner()],
            "6":['account-clock','on_press',lambda x: self.showaddresident()]
        }
    def callback(self,*args):
        # self.ids.SM.current=item
        self.ids.speed_dial.close_stack()

    def select_neighbor(self):
        neighbors_list=[]
        self.ids.neighbor_pad.clear_widgets()

        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """SELECT * FROM neighbors """,
        )
        neighbors = c.fetchall()
        for neighbor in neighbors:
            ItemConfirm.text=get_display(reshape(f'[font=bijan.ttf]{neighbor[1]}{neighbor[2]}[/font]'))
            Neighbors_Content().add_widget(ItemConfirm())

        select_neighbor_dialog=MDDialog(
            title=get_display(reshape('انتخاب کنید')),
            type='custom',
            content_cls=Neighbors_Content(),

            buttons=[
                MDRaisedButton(text='Ok')
            ]
        )
        select_neighbor_dialog.open()





        conn.commit()
        conn.close()


    # def get_fund(self,*args):
    #     incomes=0
    #     payments=0
    #
    #     conn = sqlite3.connect('database.db')
    #     # create cursor
    #     c = conn.cursor()
    #
    #     # insert into table
    #     c.execute(
    #         """ SELECT * FROM payments """,
    #     )
    #     my_payments = c.fetchall()
    #     for i in my_payments:
    #         payments+=int(i[1])
    #     c.execute(
    #         """ SELECT * FROM incomes """,
    #     )
    #     my_incomes = c.fetchall()
    #     for i in my_incomes:
    #         incomes += int(i[2])
    #
    #     funds = incomes - payments
    #     self.ids.report_total_fund.text=f'{funds}'

    def change_screen(self,screen):
        self.ids.SM.current=screen
    def on_enter(self, *args):
        self.speeddialStart()
        # self.startconnecttodatabase()
        self.Retrieve_from_neighbors_list()
        self.Retrieve_from_owners_list()
        self.Retrieve_from_income_table()
        self.Retrieve_from_payments_table()
        # self.get_fund()
        self.retrieveundonenotes()
        self.retrievedonenotes()
       # self.visiblestartbutton()
        self.get_hamsaye_number()
        self.get_owners_number()
        self.get_service_number()
        self.get_task_number()
        self.addunitnumberstomain()
        Clock.schedule_interval(self.get_time,1)





    show = 0
    def showaddnote(self,axis_y):
        if self.show == 0:
            Animation(pos_hint={'center_x': 0.5, 'center_y':float(axis_y)}).start(self.ids.adddialogbox)
            self.show = 1
        else:
            Animation(pos_hint={'center_x': 0.5, 'center_y': -1}).start(self.ids.adddialogbox)
            self.show = 0



    def showaddincome(self, *args):
        self.ids.addpad.clear_widgets()

        self.ids.adddialogbox.size_hint = (1, 0.8)
        self.ids.addpad.add_widget(AddIncome())
        self.showaddnote('0.35')
        self.callback()
    def showaddpayment(self, *args):
        self.ids.addpad.clear_widgets()

        self.ids.adddialogbox.size_hint=(1,0.7)
        self.ids.addpad.add_widget(AddPayment())

        self.showaddnote('0.35')
        self.callback()
    def showaddnotemain(self, *args):
        self.ids.addpad.clear_widgets()
        self.ids.adddialogbox.size_hint = (1, 0.6)

        self.ids.addpad.add_widget(AddNote())
        self.showaddnote(0.25)
        self.callback()
    def showaddowner(self, *args):
        self.ids.addpad.clear_widgets()
        self.ids.adddialogbox.size_hint = (1, 0.65)
        self.ids.addpad.add_widget(AddOwner())
        self.showaddnote(0.3)
        self.callback()
    def showaddresident(self, *args):
        self.ids.addpad.clear_widgets()
        self.ids.adddialogbox.size_hint = (1, 0.9)
        self.ids.addpad.add_widget(AddResident())
        self.showaddnote(0.4)
        self.callback()
    def showaddbuildingcog(self):
        self.ids.addpad.clear_widgets()
        self.ids.adddialogbox.size_hint = (1, 0.6)
        self.ids.addpad.add_widget(AddBuildingCog())
        self.showaddnote('0.25')
        self.callback()





    def addtasktodatabase(self, *args):
        if self.ids.task_title_txt.text!='' and self.ids.task_desc_txt.text!='':
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute("""INSERT INTO new_tasks VALUES (:ititle,:idescription) """,
                      {'ititle': self.ids.task_title_txt.text,
                       'idescription': self.ids.task_desc_txt.text,
                       })
            conn.commit()
            conn.close()
            self.retrieveundonenotes()

            self.ids.task_title_txt.text=''
            self.ids.task_title_input.text=''
            self.ids.task_desc_txt.text=''
            self.ids.task_desc_input.text=''
            self.on_enter()
        else:
            Mainapp().system_message('.عنوان و شرح را وارد کنید', 'باشه')


    def adddonetasktodatabase(self, *args):
        self.retrieveundonenotes()
        self.on_enter()

    def retrieveundonenotes(self, *args):
        self.ids.undonenotepad.clear_widgets()
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        tasks = c.execute(
            """ SELECT * FROM new_tasks"""
        )
        for task in tasks:
            MyNote.title = task[0]
            MyNote.description = task[1]
            self.ids.undonenotepad.add_widget(MyNote())
        conn.commit()
        conn.close()


    def retrievedonenotes(self, *args):
        self.ids.donenotepad.clear_widgets()
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        tasks = c.execute(
            """ SELECT * FROM done_tasks"""
        )
        for task in tasks:
            Donenote.title = task[0]
            Donenote.description = task[1]
            self.ids.donenotepad.add_widget(Donenote())

        conn.commit()
        conn.close()




    # Retrieve list of neighbors from neighbors table and show it in Neighbor's screen
    def Retrieve_from_neighbors_list(self):
        self.ids.neighbor_pad.clear_widgets()

        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """ SELECT * FROM neighbors """,
        )
        neighbors=c.fetchall()
        for neighbor in neighbors:
            NeighborsListReportItems.my_text=get_display(f'{neighbor[0]} {neighbor[1]}')
            NeighborsListReportItems.secondary_text=f"{neighbor[2]}"
            self.ids.neighbor_pad.add_widget(NeighborsListReportItems())

        conn.commit()
        conn.close()

        # Retrieve list of owners from owners table and show it in owner's screen
    def Retrieve_from_owners_list(self):
        self.ids.owners_pad.clear_widgets()
        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """ SELECT * FROM owners """,
        )
        owners = c.fetchall()
        for owner in owners:
            OwnersListReportItems.my_text = get_display(f'{owner[1]} {owner[0]}')
            OwnersListReportItems.secondary_text = f"{owner[2]}"
            self.ids.owners_pad.add_widget(OwnersListReportItems())

        conn.commit()
        conn.close()
    #Retrieve list of incomes from income table and show it in Income's screen
    def Retrieve_from_income_table(self):
        self.ids.income_report.clear_widgets()

        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """ SELECT * FROM incomes """,
        )
        incomes=c.fetchall()
        for income in incomes:
            IncomesListReportItems.unit_text=f'{income[0]}'
            IncomesListReportItems.date_text=f'{income[1]}'
            IncomesListReportItems.amount_text=f'{income[2]}'
            IncomesListReportItems.for_text=f'{income[3]}'

            self.ids.income_report.add_widget(IncomesListReportItems())

        conn.commit()
        conn.close()
    def Retrieve_from_payments_table(self):
        self.ids.payment_report.clear_widgets()

        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """ SELECT * FROM payments """,
        )
        pays=c.fetchall()
        for pay in pays:
            PaymentsListReportItems.date_text=f'{pay[0]}'
            PaymentsListReportItems.amount_text=f'{pay[1]}'
            PaymentsListReportItems.for_text=get_display(reshape(f'{pay[2]}'))
            PaymentsListReportItems.whom_text=f'{pay[3]}'

            self.ids.payment_report.add_widget(PaymentsListReportItems())

        conn.commit()
        conn.close()
        print('PAYMENT')
    def Retrieve_from_building_cogs_table(self):
        self.ids.building_cog_pad.clear_widgets()
        bg_colors=['Red','Blue','Purple','Amber','Teal','Pink']
        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # insert into table
        c.execute(
            """ SELECT * FROM building_cog """,
        )
        cogs=c.fetchall()
        for cog in cogs:
            BuildingCogListReportItems.name_text=get_display(reshape(f'{cog[0]}'))
            BuildingCogListReportItems.type_text=get_display(reshape(f'{cog[1]}'))
            BuildingCogListReportItems.phone_text=f'{cog[2]}'
            BuildingCogListReportItems.bg_color=choice(bg_colors)
            self.ids.building_cog_pad.add_widget(BuildingCogListReportItems())

        conn.commit()
        conn.close()
        print('NEW NEW NEW')

    # # Connect to table and insert and retrieve data from
    # def startconnecttodatabase(self):
    #     if not exists('database.db'):
    #         # connect to database
    #         import sqlite3
    #         conn = sqlite3.connect('database.db')
    #         # create cursor
    #         c = conn.cursor()
    #
    #         # create tables
    #
    #         c.execute(
    #             """ CREATE TABLE owners ("iName" text,"FamilyName" text ,"PhoneNumber" text, 'UnitNumber' text) """)
    #         # Create a table for all users this deleted neighbors are in this table
    #         c.execute(
    #             """ CREATE TABLE neighbors ("iName" text,"FamilyName" text ,"PhoneNumber" text, "DateIn" text, 'UnitNumber' text,'FamilyMember' text) """)
    #         # Create a table for all users this deleted neighbors are in this table
    #
    #         c.execute(
    #             """ CREATE TABLE deleted_neighbors ("iName" text,"FamilyName" text ,"PhoneNumber" text, "DateIn" text, 'UnitNumber' text,'FamilyMember' text) """)
    #         # Where incomes registered
    #         c.execute(
    #             """CREATE TABLE incomes ("unit_text" text,"income_date" text, "amount" text, "for_info" text)""")
    #
    #         # Where Payments registered
    #         c.execute(
    #             """CREATE TABLE payments ("payment_date" text, "amount" text, "for_info" text,"to_whom" text )""")
    #
    #         # Where Building Cogs registered
    #         c.execute(
    #             """CREATE TABLE building_cog ("name" text, "type" text, "phone" text)""")
    #         # Where Tasks are registered
    #         c.execute(
    #             """CREATE TABLE new_tasks ("title" text, "description" text)""")
    #         c.execute(
    #             """CREATE TABLE done_tasks ("title" text, "description" text)""")
    #         # Where Settings registered
    #         c.execute(
    #             """CREATE TABLE setting ("color" text, "ManagerName" text,"BuildingName" text)""")
    #         c.execute(""" INSERT INTO setting VALUES(:my_color,:manager_name,:building_name)""",
    #                   {'my_color':'Blue',
    #                    'manager_name':sm.get_screen('startscreen').ids.manager_name,
    #                    'building_name':sm.get_screen('startscreen').ids.building_name
    #
    #                    })
    #         conn.commit()
    #         conn.close()
    #     else:
    #         pass
    #



    def savesettingcolor(self,icolor,*args):
        conn = sqlite3.connect('database.db')
        #     # create cursor
        c = conn.cursor()

        # Update table
        c.execute(
            f"""UPDATE setting SET color='{icolor}' """)
        print(icolor)
        conn.commit()
        conn.close()

    #Create a new account
    def create_new_owner(self):
        if self.ids.owner_name.text and self.ids.owner_family_name.text and self.ids.owner_phone_number.text and self.ids.owner_unit_number.text !='':
            conn = sqlite3.connect('database.db')
            # create cursor
            c = conn.cursor()

            # insert into table
            c.execute(
                """INSERT INTO owners VALUES(:iName,:FamilyName,:PhoneNumber,:UnitNumber) """,
                {
                    'iName':self.ids.neighbor_name.text,
                    'FamilyName':self.ids.neighbor_family_name.text,
                    'PhoneNumber':self.ids.neighbor_phone_number.text,
                    'UnitNumber':self.ids.neighbor_unit_number.text,
                })
            conn.commit()
            conn.close()
            # Clear all the Fields:
            self.ids.owner_name.text=''
            self.ids.owner_unit_number.text=''
            self.ids.owner_family_name.text=''
            self.ids.owner_phone_number.text=''
            self.on_enter()
            # Message that function has been Done.
            Mainapp().system_message('.مالک جدید افزوده شد','عالیه')
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')

        # Create a new resident

    def create_new_account(self):
        if self.ids.neighbor_name.text and self.ids.neighbor_family_name.text and self.ids.neighbor_phone_number.text and self.ids.date_in.text and self.ids.neighbor_unit_number.text and self.ids.neighbor_family_number.text !='':
            conn = sqlite3.connect('database.db')
            # create cursor
            c = conn.cursor()

            # insert into table
            c.execute(
                """INSERT INTO neighbors VALUES(:iName,:FamilyName ,:PhoneNumber, :DateIn, :UnitNumber,:FamilyMember) """,
                {
                    'iName':self.ids.neighbor_name.text,
                    'FamilyName':self.ids.neighbor_family_name.text,
                    'PhoneNumber':self.ids.neighbor_phone_number.text,
                    'DateIn':self.ids.date_in.text,
                    'UnitNumber':self.ids.neighbor_unit_number.text,
                    'FamilyMember':self.ids.neighbor_family_number.text

                })
            c.execute(
                """INSERT INTO deleted_neighbors VALUES(:iName,:FamilyName ,:PhoneNumber, :DateIn, :UnitNumber,:FamilyMember) """,
                {
                    'iName': self.ids.neighbor_name.text,
                    'FamilyName': self.ids.neighbor_family_name.text,
                    'PhoneNumber': self.ids.neighbor_phone_number.text,
                    'DateIn': self.ids.date_in.text,
                    'UnitNumber': self.ids.neighbor_unit_number.text,
                    'FamilyMember': self.ids.neighbor_family_number.text

                })


            conn.commit()
            conn.close()
            # Clear all the Fields:
            self.ids.neighbor_name.text=''
            self.ids.neighbor_family_name.text=''
            self.ids.neighbor_phone_number.text=''
            self.ids.date_in.text=''
            self.ids.neighbor_unit_number.text=''
            self.ids.neighbor_family_number.text=''

            self.on_enter()
            # Message that function has been Done.
            Mainapp().system_message('.همسایه جدید افزوده شد','عالیه')
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')


    # Show the search payment Box:
    def showsearchpaymentbox(self, *args):
        self.ids.searchSM.current = 'searchpaymenttab'
        Animation(pos_hint={'center_y': 0.1}, d=0.2).start(self.ids.reserchbox)
    def searchPayment(self,itext):
        if self.ids.research_payment_text.text != "":
            Animation(pos_hint={'center_y': -0.3}, d=0.2).start(self.ids.reserchbox)

            conn=sqlite3.connect('database.db')
            c=conn.cursor()

            payments=c.execute(""" SELECT * FROM payments """)
            for pay in payments:
                if pay[2]==itext:
                    PaymentsListReportItems.amount_text=pay[1]
                    PaymentsListReportItems.whom_text=pay[3]
                    PaymentsListReportItems.for_text=get_display(pay[2])
                    PaymentsListReportItems.date_text=pay[0]
                    self.ids.tab_payment_report.add_widget(PaymentsListReportItems())

                else:
                    Mainapp().system_message('.موردی پیدا نشد', 'باشه')

            conn.commit()
            conn.close()
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')
    def showsearchincomebox(self, *args):

        Animation(pos_hint={'center_y': 0.1}, d=0.2).start(self.ids.reserchincomebox)
        self.ids.searchSM.current='searchincometab'

    def searchIncome(self, itext):
        if self.ids.research_income_text.text!="":
            Animation(pos_hint={'center_y': -0.3}, d=0.2).start(self.ids.reserchincomebox)

            conn = sqlite3.connect('database.db')
            c = conn.cursor()

            incomes = c.execute(""" SELECT * FROM incomes """)
            for income in incomes:
                if income[0] == itext:
                    IncomesListReportItems.unit_text=income[0]
                    IncomesListReportItems.date_text=income[1]
                    IncomesListReportItems.for_text=income[3]
                    IncomesListReportItems.amount_text=income[2]
                    self.ids.tab_income_report.add_widget(IncomesListReportItems())
                else:
                    Mainapp().system_message('.موردی پیدا نشد', 'باشه')

            conn.commit()
            conn.close()
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')




    # Search for Building Services
    def showsearchsevicebox(self, *args):

        Animation(pos_hint={'center_y': 0.1}, d=0.2).start(self.ids.searchservicebox)
        self.ids.searchSM.current='searchservicetab'

    def searchServices(self, jtext):
        if self.ids.research_service_text.text != "":
            Animation(pos_hint={'center_y': -0.3}, d=0.2).start(self.ids.searchservicebox)

            conn = sqlite3.connect('database.db')
            c = conn.cursor()

            services = c.execute(""" SELECT * FROM building_cog """)
            for service in services:
                print(service)
                if service[1] == jtext:
                    BuildingCogListReportItems.name_text=get_display(service[1])
                    BuildingCogListReportItems.phone_text=service[2]
                    BuildingCogListReportItems.type_text=get_display(service[0])

                    self.ids.tab_service_report.add_widget(BuildingCogListReportItems())
                else:
                    Mainapp().system_message('.موردی پیدا نشد', 'باشه')

            conn.commit()
            conn.close()
        else:
            Mainapp().system_message('.اطلاعات را کامل کنید', 'باشه')

class Mainapp(MDApp):

    def setcurrentthemecolor(self):
        conn = sqlite3.connect('database.db')
        # create cursor
        c = conn.cursor()

        # Update table
        theme_colors = c.execute(
            """SELECT * FROM setting  """)
        for theme in theme_colors:
            return theme[0]

        conn.commit()
        conn.close()

    font_name='bijan.ttf'
    def build(self):
        global sm
        sm = ScreenManager(transition=NoTransition())
        sm.add_widget(StartScreen(name='startscreen'))
        sm.add_widget(MyScreen(name='main'))
        if exists('database.db'):
            sm.current='main'
            if self.setcurrentthemecolor()==None:
                self.theme_cls.primary_palette='Amber'
            else:
                self.theme_cls.primary_palette =self.setcurrentthemecolor()
        else:
            sm.current='startscreen'

        return sm




    system_message_dialog=None
    def system_message(self,payam,accept):
        if not self.system_message_dialog:
            self.system_message_dialog=MDDialog(
                title=get_display(reshape(f"[font=bijan.ttf]پیام سیستم[/font]")),
                text=get_display(reshape(f"[font=bijan.ttf]{payam}[/font]")),
                buttons=[
                    MDRaisedButton(text=get_display(reshape(f"[font=bijan.ttf]{accept}[/font]")), on_press=self.close_system_message_dialog)
                ]
            )
        self.system_message_dialog.open()

    def close_system_message_dialog(self,*args):
        self.system_message_dialog.dismiss()


    def gotosearchbuildingcog(self):
        sm.get_screen('main').ids.searchSM.current = 'searchbuildingcog'





Mainapp().run()
